from flask import Flask, render_template, request, redirect, send_file,url_for,jsonify,flash
import sqlite3
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import threading
import webbrowser
import os
import sys
from datetime import datetime,timedelta
from openpyxl.styles import PatternFill
import hashlib
import json
from cryptography.fernet import Fernet
import uuid
import platform
import json
import hashlib
from flask import session,send_file
import csv
import re


# =========================================================
# 🔐 LICENSE SYSTEM
# =========================================================

# ✅ Generate once using Fernet.generate_key()
SECRET_KEY = b'3fL3rgrCK4gy_1cgQBQfaxFLDFC7ujHQkv0vwRbqezo='
fernet = Fernet(SECRET_KEY)

# Safe path (works in EXE + all PCs)
base_path = os.getenv("APPDATA") or os.path.expanduser("~")
LICENSE_FILE = os.path.join(base_path, ".sys_cache.dat")

def get_device_id():
    raw = f"{uuid.getnode()}-{platform.system()}-{platform.processor()}"
    return hashlib.sha256(raw.encode()).hexdigest()

def hash_key(key):
    return hashlib.sha256(key.encode()).hexdigest()

# 🔑 YOUR VALID KEYS
VALID_KEYS = ["KISHU-2026-PRO-91X7"]
VALID_HASHES = [hash_key(k) for k in VALID_KEYS]

def save_license(key):
    data = {
        "activated": True,
        "device": get_device_id(),
        "key_hash": hash_key(key)
    }
    encrypted = fernet.encrypt(json.dumps(data).encode())

    with open(LICENSE_FILE, "wb") as f:
        f.write(encrypted)

def is_activated():
    if not os.path.exists(LICENSE_FILE):
        return False

    try:
        with open(LICENSE_FILE, "rb") as f:
            decrypted = fernet.decrypt(f.read())
            data = json.loads(decrypted.decode())

        if not isinstance(data, dict):
            return False

        if "activated" not in data:
            return False

        if data["device"] != get_device_id():
            return False

        if data["key_hash"] not in VALID_HASHES:
            return False

        return True

    except:
        return False



app = Flask(__name__)
app.secret_key = "stock_system_2026"

DATABASE = "inventory.db"
DB = "database.db"

# Check database exists
inventory_exists = os.path.exists(DATABASE)
credit_exists = os.path.exists(DB)


# 🔒 PROTECT ALL ROUTES
@app.before_request
def check_license():
    if request.endpoint is None:
        return

    allowed_routes = ["activate", "static"]

    if request.endpoint not in allowed_routes:
        if not is_activated():
            return redirect("/activate")

# =========================================================
# 🔑 ACTIVATION ROUTE
# =========================================================
@app.route("/activate", methods=["GET", "POST"])
def activate():
    if request.method == "POST":
        key = request.form["key"]

        if hash_key(key) in VALID_HASHES:
            save_license(key)
            return redirect("/home")

        return "Invalid Key"

    return render_template("activate.html")

# =========================================================
# DATABASE
# =========================================================
# =====================================================
# DATABASE CONNECTION
# =====================================================
def get_db_connection():
    conn = sqlite3.connect(
        DATABASE,
        timeout=30,
        check_same_thread=False
    )
    conn.execute("PRAGMA foreign_keys = ON")

    conn.row_factory = sqlite3.Row
    return conn


# =====================================================
# INITIALIZE DATABASE
# =====================================================
def init_db_connection():

    conn_connection = get_db_connection()
    cur_connection = conn_connection.cursor()

    cur_connection.executescript("""

    -- CATEGORY TABLE
    CREATE TABLE IF NOT EXISTS Category (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        Cat_Name TEXT NOT NULL
    );

    -- ITEM TABLE
    CREATE TABLE IF NOT EXISTS Item (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        Item_Name TEXT NOT NULL,
        Cat_Id INTEGER,

        FOREIGN KEY (Cat_Id) REFERENCES Category(Id)
    );

    -- PRICE UPDATE TABLE
    CREATE TABLE IF NOT EXISTS Price_Update (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        Date TEXT NOT NULL
    );

    -- PRICE LIST TABLE
    CREATE TABLE IF NOT EXISTS priceList (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        ItemID INTEGER NOT NULL,
        NetPrice REAL,
        SellPrice REAL,
        MRP REAL,
        PriceUpdateID INTEGER,

        FOREIGN KEY (ItemID) REFERENCES Item(Id),
        FOREIGN KEY (PriceUpdateID) REFERENCES Price_Update(Id)
    );

    -- STOCK TABLE
    CREATE TABLE IF NOT EXISTS Stock (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        Count INTEGER NOT NULL,
        P_Id INTEGER,

        FOREIGN KEY (P_Id) REFERENCES priceList(Id)
    );

    -- STOCK IN TABLE
    CREATE TABLE IF NOT EXISTS Stock_In (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        P_Id INTEGER NOT NULL,
        Count INTEGER NOT NULL DEFAULT 0,

        FOREIGN KEY (P_Id)
        REFERENCES priceList(Id)
        ON DELETE CASCADE
    );

    -- STOCK OUT TABLE
    CREATE TABLE IF NOT EXISTS Stock_Out (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        P_Id INTEGER NOT NULL,
        Count INTEGER NOT NULL DEFAULT 0,

        FOREIGN KEY (P_Id)
        REFERENCES priceList(Id)
        ON DELETE CASCADE
    );

    -- TRANSFER TABLE
    CREATE TABLE IF NOT EXISTS Transfer (
        Id INTEGER PRIMARY KEY AUTOINCREMENT,
        Lorry_Number TEXT,
        Date TEXT
    );

    -- TRANSFER STOCK TABLE
    CREATE TABLE IF NOT EXISTS TransferStock (
        Tr_Id INTEGER,
        Sto_Id INTEGER,
        OUT_Count INTEGER,
        IN_Count INTEGER,

        FOREIGN KEY (Tr_Id) REFERENCES Transfer(Id),
        FOREIGN KEY (Sto_Id) REFERENCES Stock(Id)
    );

    """)

    conn_connection.commit()
    conn_connection.close()

    print("Database Initialized Successfully")

# Create inventory database only if not exists
if not inventory_exists:
    init_db_connection()
    print("inventory.db created")
else:
    print("inventory.db already exists - using existing database")

USER_FILE = "users.json"


# =========================
# DATABASE CONNECTION
# =========================
def get_db():
    conn = sqlite3.connect(DB, timeout=30, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        store_name TEXT,
        phone1 TEXT,
        phone2 TEXT,
        address TEXT,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER,
        type TEXT,
        amount REAL,
        date TEXT,
        due_date TEXT,
        note TEXT,
        payment_method TEXT,
        payment_reference TEXT,
        FOREIGN KEY(customer_id) REFERENCES customers(id)
    )
    """)



    cur.execute("""
CREATE TABLE IF NOT EXISTS settlements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    credit_id INTEGER,
    debit_id INTEGER,
    amount REAL
)
""")

        # OVERPAYMENT
    cur.execute("""
    CREATE TABLE IF NOT EXISTS customer_balance (
        customer_id INTEGER PRIMARY KEY,
        overpayment REAL DEFAULT 0
    )
    """)

    conn.commit()
    conn.close()


# Create credit database only if not exists
if not credit_exists:
    init_db()
    print("database.db created")
else:
    print("database.db already exists - using existing database")



# -------------------------------
# CREATE FILE IF NOT EXISTS
# -------------------------------
if not os.path.exists(USER_FILE):
    with open(USER_FILE, "w") as f:
        json.dump({}, f)


# -------------------------------
# LOAD / SAVE USERS
# -------------------------------
def load_users():
    with open(USER_FILE, "r") as f:
        return json.load(f)


def save_users(users):
    with open(USER_FILE, "w") as f:
        json.dump(users, f, indent=4)


# -------------------------------
# PASSWORD HASH
# -------------------------------
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# ===============================
# ROUTES
# ===============================

@app.route("/")
def home():
    return render_template("login.html")


# -------------------------------
# REGISTER
# -------------------------------
@app.route("/register", methods=["POST"])
def register():
    username = request.form["username"].strip()
    password = request.form["password"].strip()

    users = load_users()

    if username in users:
        return "User already exists"

    users[username] = hash_password(password)
    save_users(users)

    return redirect(url_for("home"))


# -------------------------------
# LOGIN
# -------------------------------
@app.route("/", methods=["POST"])
def login():
    username = request.form["username"].strip()
    password = hash_password(request.form["password"].strip())

    users = load_users()

    if username in users and users[username] == password:
        session["user"] = username
        return redirect("/home")

    flash("Invalid username or password. Please try again.", "danger")
    return redirect("/")



# -------------------------------
# LOGOUT
# -------------------------------
@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect("/")


# -------------------------------
# GLOBAL CHECK (LICENSE + LOGIN)
# -------------------------------
@app.before_request
def check_all():

    if request.endpoint is None:
        return

    allowed = ["login", "register", "home", "static", "activate"]

    # 🔐 License check (if you use it)
    if request.endpoint not in ["activate", "static"]:
        if 'is_activated' in globals():
            if not is_activated():
                return redirect("/activate")

    # 👤 Login check
    if request.endpoint not in allowed:
        if "user" not in session:
            return redirect("/")



# ---------------- ROUTES ----------------
@app.route("/home")
def index():

    # =========================
    # INVENTORY DB
    # =========================
    conn1 = sqlite3.connect("inventory.db")
    conn1.row_factory = sqlite3.Row

    total_items = conn1.execute("""
        SELECT COUNT(*) FROM Item
    """).fetchone()[0]

    total_updates = conn1.execute("""
        SELECT COUNT(*) FROM Price_Update
    """).fetchone()[0]

    # -------------------------
    # Recent Items
    # -------------------------
    recent_items = conn1.execute("""
        SELECT
            i.Id,
            i.Item_Name,
            c.Cat_Name,

            IFNULL(pl.SellPrice, 0) AS SellPrice,
            IFNULL(pl.NetPrice, 0) AS NetPrice,
            IFNULL(pl.MRP, 0) AS MRP,

            IFNULL(SUM(s.Count), 0) AS Stock

        FROM Item i

        LEFT JOIN Category c ON c.Id = i.Cat_Id

        LEFT JOIN PriceList pl
            ON pl.Id = (
                SELECT pl2.Id
                FROM PriceList pl2
                WHERE pl2.ItemID = i.Id
                ORDER BY pl2.Id DESC
                LIMIT 1
            )

        LEFT JOIN PriceList pl_all ON pl_all.ItemID = i.Id
        LEFT JOIN Stock s ON s.P_Id = pl_all.Id

        GROUP BY i.Id
        ORDER BY i.Id DESC
        LIMIT 10
    """).fetchall()

    conn1.close()

    # =========================
    # CREDIT DB
    # =========================
    conn2 = sqlite3.connect("database.db")
    conn2.row_factory = sqlite3.Row

    total_customers = conn2.execute("""
        SELECT COUNT(*) FROM customers
    """).fetchone()[0]

    credit_row = conn2.execute("""
        SELECT
            IFNULL(SUM(
                CASE
                    WHEN type='credit' THEN amount
                    WHEN type='debit' THEN -amount
                END
            ),0)
        FROM transactions
    """).fetchone()

    total_credit = "{:,.2f}".format(credit_row[0])

    conn2.close()

    # -------------------------
    today = datetime.now().strftime("%d %B %Y")

    return render_template(
        "home.html",
        total_items=total_items,
        total_updates=total_updates,
        total_customers=total_customers,
        total_credit=total_credit,
        recent_items=recent_items,
        today=today
    )
#---------------ItamManage ROUTES-----------------------------------

@app.route('/ItamManage', methods=['GET'])

def manage_Itam():

    conn = get_db_connection()
    items = conn.execute("SELECT Item.Id,Item.Item_Name,Category.Cat_Name,Item.Cat_ID FROM Item INNER JOIN Category ON Item.Cat_Id = Category.Id;").fetchall()
    Categorys = conn.execute("SELECT * FROM Category").fetchall()
    conn.close()

    return render_template('ItamManage.html',items=items,Categorys=Categorys)

@app.route('/add_item', methods=['POST'])
def add_item():

    name = request.form['name']
    catID = request.form['catID']

    conn = get_db_connection()
    cursor = conn.cursor()

    try:

        # ==================================
        # INSERT NEW ITEM
        # ==================================
        cursor.execute("""

            INSERT INTO Item
            (
                Item_Name,
                Cat_Id
            )

            VALUES (?,?)

        """,(name,catID))

        # NEW ITEM ID
        item_id = cursor.lastrowid

        # ==================================
        # GET LATEST PRICE UPDATE ID
        # ==================================
        latest_price = cursor.execute("""

            SELECT Id
            FROM Price_Update
            ORDER BY Id DESC
            LIMIT 1

        """).fetchone()

        # ==================================
        # AUTO CREATE PRICELIST RECORD
        # ==================================
        if latest_price:

            price_update_id = latest_price["Id"]

            cursor.execute("""

                INSERT INTO priceList
                (
                    ItemID,
                    PriceUpdateID,
                    NetPrice,
                    SellPrice,
                    MRP
                )

                VALUES (?,?,?,?,?)

            """,(

                item_id,
                price_update_id,
                0,
                0,
                0

            ))

        conn.commit()

    except Exception as e:

        conn.rollback()

        print("ADD ITEM ERROR :",e)

    finally:

        conn.close()

    return redirect('/ItamManage')


@app.route('/update_item', methods=['POST'])
def update_item():

    item_id = request.form['itemId']
    name = request.form['name']
    catID = request.form['catID']

    conn = get_db_connection()

    try:

        conn.execute("""

            UPDATE Item

            SET
                Item_Name=?,
                Cat_Id=?

            WHERE Id=?

        """,(

            name,
            catID,
            item_id

        ))

        conn.commit()

    except Exception as e:

        conn.rollback()

        print("UPDATE ITEM ERROR :",e)

    finally:

        conn.close()

    return redirect('/ItamManage')

@app.route('/get_delete_details/<int:item_id>')
def get_delete_details(item_id):

    conn = get_db_connection()
    cursor = conn.cursor()

    data = cursor.execute("""

        SELECT

            pu.Date,
            IFNULL(s.Count,0) AS stock,

            pl.NetPrice,
            pl.SellPrice,
            pl.MRP

        FROM priceList pl

        LEFT JOIN Price_Update pu
            ON pu.Id = pl.PriceUpdateID

        LEFT JOIN Stock s
            ON s.P_Id = pl.Id

        WHERE pl.ItemID=?

        ORDER BY pu.Id DESC

    """,(item_id,)).fetchall()

    conn.close()

    result=[]

    for row in data:

        result.append({

            "date":row["Date"],
            "stock":row["stock"],
            "net":row["NetPrice"],
            "sell":row["SellPrice"],
            "mrp":row["MRP"]

        })

    return jsonify(result)



@app.route('/deleteItem/<int:item_id>')
def delete_item(item_id):

    conn = get_db_connection()
    cursor = conn.cursor()

    try:

        # -----------------------------
        # GET ITEM NAME
        # -----------------------------
        item = cursor.execute("""
            SELECT Item_Name
            FROM Item
            WHERE Id=?
        """, (item_id,)).fetchone()

        if not item:
            flash("Item not found", "danger")
            return redirect('/ItamManage')

        # -----------------------------
        # GET PRICELIST IDS
        # -----------------------------
        price_ids = cursor.execute("""
            SELECT Id
            FROM priceList
            WHERE ItemID=?
        """, (item_id,)).fetchall()

        p_ids = [str(x["Id"]) for x in price_ids]

        stock_count = 0

        if p_ids:

            placeholders = ",".join(["?"] * len(p_ids))

            stock_count = cursor.execute(f"""
                SELECT COUNT(*)
                FROM Stock
                WHERE P_Id IN ({placeholders})
            """, p_ids).fetchone()[0]

        # -----------------------------
        # DELETE RELATED DATA
        # -----------------------------
        if p_ids:

            placeholders = ",".join(["?"] * len(p_ids))

            cursor.execute(f"""
                DELETE FROM Stock_In
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            cursor.execute(f"""
                DELETE FROM Stock_Out
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            cursor.execute(f"""
                DELETE FROM TransferStock
                WHERE Sto_Id IN (
                    SELECT Id
                    FROM Stock
                    WHERE P_Id IN ({placeholders})
                )
            """, p_ids)

            cursor.execute(f"""
                DELETE FROM Stock
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            cursor.execute("""
                DELETE FROM priceList
                WHERE ItemID=?
            """, (item_id,))

        # -----------------------------
        # DELETE ITEM
        # -----------------------------
        cursor.execute("""
            DELETE FROM Item
            WHERE Id=?
        """, (item_id,))

        conn.commit()

        flash(
            f"{item['Item_Name']} deleted successfully "
            f"(Removed related stock records: {stock_count})",
            "success"
        )

    except Exception as e:

        conn.rollback()

        flash(
            f"Delete Failed : {str(e)}",
            "danger"
        )

    finally:
        conn.close()

    return redirect('/ItamManage')



# ==================================================
# CATEGORY CRUD
# ==================================================

# ADD CATEGORY
@app.route("/add_Category", methods=["POST"])
def add_Category():
    Cat_name = request.form["Cat_name"]

    conn = get_db_connection()
    conn.execute(
        "INSERT INTO Category (Cat_Name) VALUES (?)",
        (Cat_name,)
    )
    conn.commit()
    conn.close()

    flash("Category Added")
    return redirect("/ItamManage")


# UPDATE CATEGORY
@app.route("/update_Category", methods=["POST"])
def update_Category():
    catId = request.form["catId"]
    Cat_name = request.form["Cat_name"]

    conn = get_db_connection()
    conn.execute("""
        UPDATE Category
        SET Cat_Name=?
        WHERE Id=?
    """, (Cat_name, catId))
    conn.commit()
    conn.close()

    flash("Category Updated")
    return redirect("/ItamManage")


# DELETE CATEGORY + DELETE ITEMS UNDER IT
# DELETE CATEGORY + DELETE ALL ITEMS AND RELATED DATA UNDER IT
@app.route("/deleteCategory/<int:id>")
def deleteCategory(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get category name
        category = cursor.execute("SELECT Cat_Name FROM Category WHERE Id = ?", (id,)).fetchone()

        if not category:
            flash("Category not found", "danger")
            return redirect("/ItamManage")

        # Get all items in this category
        items = cursor.execute("SELECT Id FROM Item WHERE Cat_Id = ?", (id,)).fetchall()

        for item in items:
            item_id = item["Id"]

            # Get all PriceList IDs for this item
            price_list_ids = cursor.execute("SELECT Id FROM priceList WHERE ItemID = ?", (item_id,)).fetchall()

            for pl in price_list_ids:
                p_id = pl["Id"]
                # Delete from Stock_In
                cursor.execute("DELETE FROM Stock_In WHERE P_Id = ?", (p_id,))
                # Delete from Stock_Out
                cursor.execute("DELETE FROM Stock_Out WHERE P_Id = ?", (p_id,))
                # Delete from Stock
                cursor.execute("DELETE FROM Stock WHERE P_Id = ?", (p_id,))

            # Delete from priceList
            cursor.execute("DELETE FROM priceList WHERE ItemID = ?", (item_id,))

        # Delete all items in category
        cursor.execute("DELETE FROM Item WHERE Cat_Id = ?", (id,))

        # Delete category
        cursor.execute("DELETE FROM Category WHERE Id = ?", (id,))

        conn.commit()
        flash(f'Category "{category["Cat_Name"]}" and all associated items have been deleted successfully', "success")

    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect("/ItamManage")


#===================================================================================

#---------------QuaterManage ROUTES-----------------------------------------

# =====================================================
# PRICE UPDATE MANAGE PAGE
# =====================================================
@app.route('/PriceUpdateManage')
def PriceUpdateManage():
    from datetime import datetime  # Add this import at top or here

    conn = get_db_connection()

    PriceUpdates = conn.execute("""
        SELECT *
        FROM Price_Update
        ORDER BY Id DESC
    """).fetchall()

    conn.close()

    return render_template(
        'priceUpdateManage.html',
        PriceUpdates=PriceUpdates,
        datetime=datetime  # Pass datetime to template
    )


# =====================================================
# ADD NEW PRICE UPDATE
# =====================================================
@app.route('/PriceUpdate', methods=['POST'])
def PriceUpdate():

    try:
        U_Date = request.form['upDate'].strip()

        if not U_Date:
            flash("Please select a date", "danger")
            return redirect(url_for('PriceUpdateManage'))

        # Convert date
        date_obj = datetime.strptime(U_Date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %B %Y")

        conn = get_db_connection()

        conn.execute("""
            INSERT INTO Price_Update (Date)
            VALUES (?)
        """, (formatted_date,))

        conn.commit()
        conn.close()

        flash("Price Update Added Successfully", "success")

    except Exception as e:
        flash(f"Error : {str(e)}", "danger")

    return redirect(url_for('PriceUpdateManage'))


# =====================================================
# DELETE PRICE UPDATE
# =====================================================
@app.route('/deletePriceUpdate/<int:price_update_id>')
def delete_price_update(price_update_id):
    # This route now just checks and returns JSON for the modal
    conn = get_db_connection()
    cursor = conn.cursor()

    # Check if there's any stock linked to this price update
    stock_info = cursor.execute("""
        SELECT
            i.Id as item_id,
            i.Item_Name,
            s.Count as stock_quantity,
            pl.NetPrice,
            pl.SellPrice,
            pl.MRP
        FROM PriceList pl
        JOIN Stock s ON s.P_Id = pl.Id
        JOIN Item i ON i.Id = pl.ItemID
        WHERE pl.PriceUpdateID = ?
        AND s.Count > 0
        ORDER BY i.Item_Name
    """, (price_update_id,)).fetchall()

    # Get price update date
    price_update = cursor.execute("""
        SELECT Date, Id FROM Price_Update WHERE Id = ?
    """, (price_update_id,)).fetchone()

    conn.close()

    if stock_info and len(stock_info) > 0:
        # Return stock info as JSON for the modal
        items_list = []
        for item in stock_info:
            items_list.append({
                'item_id': item['item_id'],
                'item_name': item['Item_Name'],
                'stock_quantity': item['stock_quantity'],
                'net_price': float(item['NetPrice']) if item['NetPrice'] else 0,
                'sell_price': float(item['SellPrice']) if item['SellPrice'] else 0,
                'mrp': float(item['MRP']) if item['MRP'] else 0
            })

        return jsonify({
            'has_stock': True,
            'stock_count': len(stock_info),
            'price_update_id': price_update_id,
            'price_update_date': price_update['Date'] if price_update else 'Unknown',
            'items': items_list
        })

    # No stock found - delete directly
    return jsonify({
        'has_stock': False,
        'price_update_id': price_update_id
    })


@app.route('/force_delete_price_update', methods=['POST'])
def force_delete_price_update():
    data = request.get_json()
    price_update_id = data.get('price_update_id')

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get all PriceList IDs for this price update
        price_list_ids = cursor.execute("""
            SELECT Id FROM PriceList WHERE PriceUpdateID = ?
        """, (price_update_id,)).fetchall()

        if price_list_ids:
            p_ids = [str(row['Id']) for row in price_list_ids]
            placeholders = ",".join(["?"] * len(p_ids))

            # Delete from Stock_In
            cursor.execute(f"""
                DELETE FROM Stock_In
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            # Delete from Stock_Out
            cursor.execute(f"""
                DELETE FROM Stock_Out
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            # Delete from TransferStock
            cursor.execute(f"""
                DELETE FROM TransferStock
                WHERE Sto_Id IN (
                    SELECT Id FROM Stock WHERE P_Id IN ({placeholders})
                )
            """, p_ids)

            # Delete from Stock
            cursor.execute(f"""
                DELETE FROM Stock
                WHERE P_Id IN ({placeholders})
            """, p_ids)

            # Delete from PriceList
            cursor.execute(f"""
                DELETE FROM PriceList
                WHERE PriceUpdateID = ?
            """, (price_update_id,))

        # Delete parent record
        cursor.execute("""
            DELETE FROM Price_Update
            WHERE Id = ?
        """, (price_update_id,))

        conn.commit()

        return jsonify({
            'success': True,
            'message': f'Price Update #{price_update_id} and ALL associated stock have been deleted successfully'
        })

    except Exception as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Delete Failed : {str(e)}'
        })
    finally:
        conn.close()

# =====================================================
# PRICE LIST PAGE
# =====================================================
@app.route('/priceList/<int:PriceUpdateID>')
def priceList(PriceUpdateID):

    conn = get_db_connection()

    # All Items
    items = conn.execute("""
        SELECT
            Item.Id,
            Item.Item_Name,
            Item.Cat_ID
        FROM Item
        ORDER BY Item_Name ASC
    """).fetchall()

    # Categories
    Categorys = conn.execute("""
        SELECT *
        FROM Category
        ORDER BY Cat_Name ASC
    """).fetchall()

    # Saved Prices
    prices_raw = conn.execute("""
        SELECT
            ItemID,
            NetPrice,
            SellPrice,
            MRP
        FROM PriceList
        WHERE PriceUpdateID = ?
    """, (PriceUpdateID,)).fetchall()

    conn.close()

    # Convert row object to dict
    prices = [dict(row) for row in prices_raw]

    return render_template(
        'priceList.html',
        items=items,
        Categorys=Categorys,
        PriceUpdateID=PriceUpdateID,
        prices=prices
    )


# =====================================================
# SAVE PRICE LIST
# =====================================================
@app.route('/save_prices/<int:PriceUpdateID>', methods=['POST'])
def add_price(PriceUpdateID):
    def to_float(value):
        """Comma, Rs. ඉවත් කර float එකක් බවට පරිවර්තනය කරයි"""
        if not value:
            return 0.0
        cleaned = str(value).replace(',', '').replace('Rs.', '').replace('Rs', '').strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        item_ids = request.form.getlist('item_id[]')
        net_prices = request.form.getlist('net_price[]')
        sell_prices = request.form.getlist('sell_price[]')
        mrps = request.form.getlist('mrp[]')

        updated_count = 0
        inserted_count = 0

        for i in range(len(item_ids)):
            item_id = item_ids[i]
            net = to_float(net_prices[i])
            sell = to_float(sell_prices[i])
            mrp = to_float(mrps[i])

            # මෙම PriceUpdateID එකට දැනටමත් price record එකක් තියෙනවද?
            existing = cursor.execute("""
                SELECT Id FROM PriceList
                WHERE ItemID = ? AND PriceUpdateID = ?
            """, (item_id, PriceUpdateID)).fetchone()

            if existing:
                # Update existing row (foreign keys වලට හානියක් නැහැ)
                cursor.execute("""
                    UPDATE PriceList
                    SET NetPrice = ?, SellPrice = ?, MRP = ?
                    WHERE Id = ?
                """, (net, sell, mrp, existing["Id"]))
                updated_count += 1
            else:
                # නව row එකක් insert කරන්න
                cursor.execute("""
                    INSERT INTO PriceList (PriceUpdateID, ItemID, NetPrice, SellPrice, MRP)
                    VALUES (?, ?, ?, ?, ?)
                """, (PriceUpdateID, item_id, net, sell, mrp))
                inserted_count += 1

        conn.commit()
        flash(f"Prices saved: {updated_count} updated, {inserted_count} newly added.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Save Failed: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for('priceList', PriceUpdateID=PriceUpdateID))
#==================================Stock===============================================================

@app.route("/stock")
def stock():
    conn = get_db_connection()

    categories = conn.execute("SELECT * FROM Category").fetchall()
    price_updates = conn.execute("SELECT * FROM Price_Update ORDER BY Id DESC").fetchall()

    data = conn.execute("""
        SELECT
            i.Id as item_id,
            i.Item_Name,
            c.Cat_Name,
            pu.Id as price_update_id,

            IFNULL(s.Count, 0) as stock,

            IFNULL(pl.NetPrice, 0) as net,
            IFNULL(pl.SellPrice, 0) as sell,
            IFNULL(pl.MRP, 0) as mrp

        FROM Item i

        JOIN Category c ON c.Id = i.Cat_Id

        CROSS JOIN Price_Update pu

        LEFT JOIN PriceList pl
            ON pl.ItemID = i.Id
            AND pl.PriceUpdateID = pu.Id

        LEFT JOIN Stock s
            ON s.P_Id = pl.Id

        ORDER BY i.Item_Name
    """).fetchall()

    conn.close()

    # Structure data
    result = {}

    for row in data:
        item_id = row["item_id"]

        if item_id not in result:
            result[item_id] = {
                "name": row["Item_Name"],
                "category": row["Cat_Name"],
                "prices": {}
            }

        result[item_id]["prices"][row["price_update_id"]] = {
            "stock": row["stock"],
            "net": row["net"],
            "sell": row["sell"],
            "mrp": row["mrp"]
        }

    return render_template(
        "stock.html",
        Categorys=categories,
        PriceUpdates=price_updates,
        items=result
    )

# ---------------- LOAD ITEMS BY CATEGORY ----------------
@app.route("/get_items/<int:cat_id>")
def get_items(cat_id):
    conn = get_db_connection()

    items = conn.execute(
        "SELECT Id, Item_Name FROM Item WHERE Cat_Id=?",
        (cat_id,)
    ).fetchall()

    conn.close()

    return jsonify([
        {"id": row["Id"], "name": row["Item_Name"]}
        for row in items
    ])


# ---------------- LOAD PRICE ----------------
@app.route("/get_price/<int:item_id>/<int:update_id>")
def get_price(item_id, update_id):
    conn = get_db_connection()

    price = conn.execute("""
        SELECT Id, NetPrice, SellPrice, MRP
        FROM priceList
        WHERE ItemID=? AND PriceUpdateID=?
        LIMIT 1
    """, (item_id, update_id)).fetchone()

    conn.close()

    if price:
        return jsonify({
            "status": True,
            "p_id": price["Id"],
            "net": price["NetPrice"],
            "sell": price["SellPrice"],
            "mrp": price["MRP"]
        })

    return jsonify({"status": False})


# ---------------- SAVE STOCK ----------------
# =====================================================
# ADD STOCK
# If P_Id exists in Stock -> add quantity
# Else insert new row
# Also insert history into Stock_In
# =====================================================
@app.route("/add_stock", methods=["POST"])
def add_stock():

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        item_id = request.form["item_id"]
        update_id = request.form["price_update_id"]
        qty = int(request.form["qty"])

        # -----------------------------------
        # Find P_Id from PriceList
        # -----------------------------------
        price = cursor.execute("""
            SELECT Id
            FROM PriceList
            WHERE ItemID = ?
            AND PriceUpdateID = ?
            LIMIT 1
        """, (item_id, update_id)).fetchone()

        if not price:
            flash("Price List record not found", "danger")
            return redirect(url_for("stock"))

        p_id = price["Id"]

        # -----------------------------------
        # Check Stock Exists
        # -----------------------------------
        old_stock = cursor.execute("""
            SELECT Id, Count
            FROM Stock
            WHERE P_Id = ?
            LIMIT 1
        """, (p_id,)).fetchone()

        if old_stock:
            # Update existing stock
            new_qty = old_stock["Count"] + qty

            cursor.execute("""
                UPDATE Stock
                SET Count = ?
                WHERE P_Id = ?
            """, (new_qty, p_id))

        else:
            # Insert new stock row
            cursor.execute("""
                INSERT INTO Stock
                (
                    P_Id,
                    Count
                )
                VALUES (?, ?)
            """, (
                p_id,
                qty
            ))

        # -----------------------------------
        # Insert Stock In History
        # -----------------------------------
        cursor.execute("""
            INSERT INTO Stock_In
            (
                P_Id,
                Count
            )
            VALUES (?, ?)
        """, (
            p_id,
            qty
        ))

        conn.commit()

        flash("Stock Added Successfully", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error : {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("stock"))



#---------------------------------------------------

@app.route("/multi_stock")
def multi_stock():

    conn = get_db_connection()
    name = request.args.get("name")

    Categorys = conn.execute("SELECT * FROM Category").fetchall()
    PriceUpdates = conn.execute("SELECT * FROM Price_Update ORDER BY Id DESC").fetchall()

    conn.close()

    return render_template(
        "multi_stock.html",
        Categorys=Categorys,
        PriceUpdates=PriceUpdates,name=name
    )


@app.route("/save_multi_stock", methods=["POST"])
def save_multi_stock():

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        item_ids = request.form.getlist("item_id[]")
        update_ids = request.form.getlist("price_update_id[]")
        qtys = request.form.getlist("qty[]")

        for i in range(len(item_ids)):

            item_id = item_ids[i]
            update_id = update_ids[i]
            qty = int(qtys[i])

            # Get price
            price = cursor.execute("""
                SELECT Id, NetPrice, SellPrice, MRP
                FROM PriceList
                WHERE ItemID = ?
                AND PriceUpdateID = ?
                LIMIT 1
            """, (item_id, update_id)).fetchone()

            # ❌ Skip if no price OR NULL values
            if not price or price["NetPrice"] is None or price["SellPrice"] is None or price["MRP"] is None:
                continue

            p_id = price["Id"]

            # Stock logic (same as before)
            stock = cursor.execute("""
                SELECT Count FROM Stock WHERE P_Id = ?
            """, (p_id,)).fetchone()

            if stock:
                new_qty = stock["Count"] + qty

                cursor.execute("""
                    UPDATE Stock SET Count = ? WHERE P_Id = ?
                """, (new_qty, p_id))

            else:
                cursor.execute("""
                    INSERT INTO Stock (P_Id, Count) VALUES (?, ?)
                """, (p_id, qty))

            # History
            cursor.execute("""
                INSERT INTO Stock_In (P_Id, Count)
                VALUES (?, ?)
            """, (p_id, qty))

        conn.commit()
        flash("Multiple Stock Saved Successfully", "success")

    except Exception as e:
        conn.rollback()
        flash(str(e), "danger")

    finally:
        conn.close()

    return redirect(url_for("multi_stock"))


@app.route("/get_multi_price/<int:item_id>/<int:update_id>")
def get_multi_price(item_id, update_id):

    conn = get_db_connection()

    price = conn.execute("""
        SELECT Id, NetPrice, SellPrice, MRP
        FROM PriceList
        WHERE ItemID=? AND PriceUpdateID=?
        LIMIT 1
    """, (item_id, update_id)).fetchone()

    conn.close()

    if price:
        return jsonify({
            "status": True,
            "net": price["NetPrice"],
            "sell": price["SellPrice"],
            "mrp": price["MRP"]
        })

    return jsonify({"status": False})





#-----------------------------------------------------------------------------

@app.route("/send_stock")
def send_stock():

    conn = get_db_connection()

    Categorys = conn.execute("SELECT * FROM Category").fetchall()
    PriceUpdates = conn.execute("SELECT * FROM Price_Update ORDER BY Id DESC").fetchall()

    conn.close()

    return render_template(
        "send_stock.html",
        Categorys=Categorys,
        PriceUpdates=PriceUpdates
    )

@app.route("/get_send_multi_price/<int:item_id>/<int:update_id>")
def get_send_multi_price(item_id, update_id):

    conn = get_db_connection()
    cursor = conn.cursor()

    price = cursor.execute("""
        SELECT Id, NetPrice, SellPrice, MRP
        FROM PriceList
        WHERE ItemID=? AND PriceUpdateID=?
        LIMIT 1
    """, (item_id, update_id)).fetchone()

    if not price:
        conn.close()
        return jsonify({"status": False})

    p_id = price["Id"]

    stock = cursor.execute("""
        SELECT Count FROM Stock WHERE P_Id = ?
    """, (p_id,)).fetchone()

    conn.close()

    return jsonify({
        "status": True,
        "net": price["NetPrice"],
        "sell": price["SellPrice"],
        "mrp": price["MRP"],
        "stok": stock["Count"] if stock else 0
    })


@app.route("/save_send_stock", methods=["POST"])
def save_send_stock():

    conn = get_db_connection()
    cursor = conn.cursor()

    saved = 0
    skipped = 0

    try:
        item_ids = request.form.getlist("item_id[]")
        update_ids = request.form.getlist("price_update_id[]")
        qtys = request.form.getlist("qty[]")

        for i in range(len(item_ids)):

            item_id = item_ids[i]
            update_id = update_ids[i]
            qty = int(qtys[i])

            price = cursor.execute("""
                SELECT Id, NetPrice, SellPrice, MRP
                FROM PriceList
                WHERE ItemID=? AND PriceUpdateID=?
                LIMIT 1
            """, (item_id, update_id)).fetchone()

            if not price or price["NetPrice"] is None or price["SellPrice"] is None or price["MRP"] is None:
                skipped += 1
                continue

            p_id = price["Id"]

            stock = cursor.execute("""
                SELECT Count FROM Stock WHERE P_Id = ?
            """, (p_id,)).fetchone()

            if stock:
                new_qty = stock["Count"] - qty

                cursor.execute("""
                    UPDATE Stock SET Count = ? WHERE P_Id = ?
                """, (new_qty, p_id))

            cursor.execute("""
                INSERT INTO Stock_Out (P_Id, Count)
                VALUES (?, ?)
            """, (p_id, qty))

            saved += 1

        conn.commit()
        flash(f"{saved} rows saved, {skipped} rows skipped", "info")

    except Exception as e:
        conn.rollback()
        flash(str(e), "danger")

    finally:
        conn.close()

    return redirect(url_for("send_stock"))


@app.route('/export_excel_stock')
def export_excel_stock():

    conn = get_db_connection()
    cur = conn.cursor()

    # LOAD DATA
    price_updates = cur.execute("""
        SELECT * FROM Price_Update ORDER BY Id DESC
    """).fetchall()

    items = cur.execute("""
        SELECT Item.Id, Item.Item_Name, Category.Cat_Name
        FROM Item
        LEFT JOIN Category ON Category.Id = Item.Cat_Id
        ORDER BY Category.Cat_Name, Item.Item_Name
    """).fetchall()

    # EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    # COLORS
    dark_blue = "1F4E78"
    green = "2E7D32"
    gold = "FFD966"
    gray = "D9E1F2"
    white = "FFFFFF"
    cyan = "D9EAF7"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ✅ UPDATED COLUMN COUNT
    total_cols = 5 + (len(price_updates) * 6)

    # TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "KISHUDHA EXPERT DISTRIBUTION"
    ws["A1"].font = Font(size=22, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = "STOCK MANAGEMENT REPORT"
    ws["A2"].font = Font(size=14, bold=True, color=white)
    ws["A2"].fill = PatternFill("solid", fgColor=green)
    ws["A2"].alignment = Alignment(horizontal="center")

    # HEADER
    ws.merge_cells("A4:A5")
    ws.merge_cells("B4:B5")
    ws.merge_cells("C4:C5")
    ws.merge_cells("D4:D5")
    ws.merge_cells("E4:E5")

    ws["A4"] = "Category"
    ws["B4"] = "Item Name"
    ws["C4"] = "Total Stock"
    ws["D4"] = "Total Net Value"
    ws["E4"] = "Total Sell Value"

    for x in ["A4", "B4", "C4", "D4", "E4"]:
        ws[x].fill = PatternFill("solid", fgColor=gold)
        ws[x].font = Font(bold=True)
        ws[x].alignment = Alignment(horizontal="center", vertical="center")
        ws[x].border = thin

    col = 6

    # DYNAMIC HEADER
    for pu in price_updates:

        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 5)

        ws.cell(row=4, column=col).value = pu["Date"]
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=dark_blue)
        ws.cell(row=4, column=col).font = Font(color=white, bold=True)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")

        names = ["Stock", "Net Price", "Net Value", "Sell Price", "Sell Value", "MRP"]

        for i in range(6):
            d = ws.cell(row=5, column=col + i)
            d.value = names[i]
            d.fill = PatternFill("solid", fgColor=gray)
            d.font = Font(bold=True)
            d.alignment = Alignment(horizontal="center")
            d.border = thin

        col += 6

    # DATA
    row_no = 6

    for item in items:

        ws.cell(row=row_no, column=1).value = item["Cat_Name"]
        ws.cell(row=row_no, column=2).value = item["Item_Name"]

        total_stock = 0
        total_net_val = 0
        total_sell_val = 0

        col = 6

        for pu in price_updates:

            price = cur.execute("""
                SELECT Id, NetPrice, SellPrice, MRP
                FROM PriceList
                WHERE ItemID=? AND PriceUpdateID=?
                LIMIT 1
            """, (item["Id"], pu["Id"])).fetchone()

            stock_qty = 0
            net = 0
            sell = 0
            mrp = 0

            if price:
                net = price["NetPrice"] or 0
                sell = price["SellPrice"] or 0
                mrp = price["MRP"] or 0

                stock = cur.execute("""
                    SELECT Count FROM Stock WHERE P_Id=? LIMIT 1
                """, (price["Id"],)).fetchone()

                if stock:
                    stock_qty = stock["Count"]

            # ✅ CALCULATIONS
            net_val = net * stock_qty
            sell_val = sell * stock_qty

            total_stock += stock_qty
            total_net_val += net_val
            total_sell_val += sell_val

            values = [stock_qty, net, net_val, sell, sell_val, mrp]

            for val in values:
                cell = ws.cell(row=row_no, column=col)
                cell.value = val
                cell.border = thin
                col += 1

        # TOTALS
        ws.cell(row=row_no, column=3).value = total_stock
        ws.cell(row=row_no, column=4).value = total_net_val
        ws.cell(row=row_no, column=5).value = total_sell_val

        row_no += 1

    # AUTO WIDTH
    for col_num in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    ws.freeze_panes = "F6"

    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    file_path = os.path.join(downloads, "Stock_Report.xlsx")
    
    wb.save(file_path)
    
    conn.close()
    
    return jsonify({
        "success": True,
        "message": "Excel Report Exported Successfully!",
        "path": file_path
    })


#===============Credit app ==============================================================================================================================================

# =========================
# HELPER FUNCTIONS
# =========================

@app.template_filter('currency')
def currency_format(value):
    try:
        return "{:,.2f}".format(float(value))
    except:
        return "0.00"


def validate_phone(phone):
    return re.match(r'^(07\d{8})$', phone)


def recalculate_customer(customer_id):
    conn = get_db()
    cur = conn.cursor()

    # ❌ reset everything
    cur.execute("DELETE FROM settlements WHERE credit_id IN (SELECT id FROM transactions WHERE customer_id=?)", (customer_id,))
    cur.execute("DELETE FROM settlements WHERE debit_id IN (SELECT id FROM transactions WHERE customer_id=?)", (customer_id,))

    cur.execute("""
        INSERT OR REPLACE INTO customer_balance (customer_id, overpayment)
        VALUES (?, 0)
    """, (customer_id,))

    # =========================
    # GET ALL CREDITS
    # =========================
    credits = cur.execute("""
        SELECT * FROM transactions
        WHERE customer_id=? AND type='credit'
        ORDER BY date ASC
    """, (customer_id,)).fetchall()

    # =========================
    # GET ALL DEBITS
    # =========================
    debits = cur.execute("""
        SELECT * FROM transactions
        WHERE customer_id=? AND type='debit'
        ORDER BY date ASC
    """, (customer_id,)).fetchall()

    # convert credits to dict
    credit_map = {}
    for c in credits:
        credit_map[c["id"]] = {
            "amount": c["amount"],
            "remaining": c["amount"]
        }

    # =========================
    # APPLY DEBITS
    # =========================
    for d in debits:

        remaining = d["amount"]

        for cid in credit_map:
            if remaining <= 0:
                break

            c = credit_map[cid]

            if c["remaining"] <= 0:
                continue

            settle = min(c["remaining"], remaining)

            cur.execute("""
                INSERT INTO settlements (credit_id, debit_id, amount)
                VALUES (?, ?, ?)
            """, (cid, d["id"], settle))

            c["remaining"] -= settle
            remaining -= settle

        # =========================
        # OVERPAYMENT
        # =========================
        if remaining > 0:
            cur.execute("""
                UPDATE customer_balance
                SET overpayment = overpayment + ?
                WHERE customer_id=?
            """, (remaining, customer_id))

    conn.commit()
    conn.close()


# =========================
# OVERPAYMENT GET
# =========================
def get_overpayment(conn, customer_id):
    cur = conn.cursor()
    row = cur.execute("""
        SELECT overpayment FROM customer_balance WHERE customer_id=?
    """, (customer_id,)).fetchone()
    return row["overpayment"] if row else 0



def calculate_balance(customer_id):
    conn = get_db()
    cur = conn.cursor()

    credit = cur.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE customer_id=? AND type='credit'",
        (customer_id,)
    ).fetchone()[0]

    debit = cur.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE customer_id=? AND type='debit'",
        (customer_id,)
    ).fetchone()[0]

    conn.close()
    return credit - debit


def get_status(balance, due_date):
    if balance == 0:
        return "Paid"

    if not due_date:
        return "Active"

    due = datetime.strptime(due_date, "%Y-%m-%d").date()
    today = datetime.today().date()

    if due < today:
        return "Expired"
    elif due == today + timedelta(days=1):
        return "Due Tomorrow"
    else:
        return "Active"



# =========================
# DASHBOARD FIXED
# =========================
@app.route("/credits")
def dashboard():
    conn = get_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()


    today = datetime.today().date()
    search = request.args.get("search", "").lower()
    status_filter = request.args.get("status", "")

    # =========================
    # GET DATA
    # =========================
    rows = cur.execute("""
        SELECT
            t.id,
            t.amount,
            t.date,
            t.due_date,
            c.name
        FROM transactions t
        JOIN customers c ON c.id = t.customer_id
        WHERE t.type = 'credit'
        ORDER BY t.date DESC
    """).fetchall()

    data = []

    total_credit = 0
    total_debit = 0
    total_balance = 0
    total_overdue = 0

    for r in rows:

        settled = cur.execute("""
            SELECT COALESCE(SUM(amount),0)
            FROM settlements
            WHERE credit_id=?
        """, (r["id"],)).fetchone()[0]

        balance = r["amount"] - settled

        # STATUS
        if balance == 0:
            status = "Paid"
        elif r["due_date"]:
            due = datetime.strptime(r["due_date"], "%Y-%m-%d").date()

            if due < today:
                status = "Expired"
            elif due == today + timedelta(days=1):
                status = "Due Tomorrow"
            else:
                status = "Active"
        else:
            status = "Active"

        # =========================
        # 🔍 FILTERS APPLY
        # =========================

        # Search filter
        if search and search not in r["name"].lower():
            continue

        # Status filter
        if status_filter and status != status_filter:
            continue

        # =========================
        # ADD DATA
        # =========================
        data.append({
            "id": r["id"],
            "customer": r["name"],
            "date": r["date"],
            "credit": r["amount"],
            "debit": settled,
            "balance": balance,
            "due": r["due_date"],
            "status": status
        })

        # =========================
        # KPI (ONLY FILTERED DATA)
        # =========================
        total_credit += r["amount"]
        total_debit += settled
        total_balance += balance

        if status == "Expired":
            total_overdue += balance


    conn.close()

    return render_template(
        "dashboard.html",
        data=data,
        total_credit=total_credit,
        total_debit=total_debit,
        total_balance=total_balance,
        total_overdue=total_overdue
    )

# =========================
# CUSTOMERS
# =========================
@app.route("/customers")
def customers():
    conn = get_db()
    cur = conn.cursor()

    search = request.args.get("search", "")

    rows = cur.execute("SELECT * FROM customers ORDER BY name ASC").fetchall()

    data = []
    for c in rows:
        balance = calculate_balance(c["id"])

        if search and search.lower() not in c["name"].lower():
            continue

        data.append({
            "id": c["id"],
            "name": c["name"],
            "store": c["store_name"],
            "balance": balance
        })

    conn.close()
    return render_template("customers.html", customers=data)


@app.route("/add_customer", methods=["GET", "POST"])
def add_customer():

    if request.method == "POST":
        name = request.form["name"].strip()
        store = request.form["store"]
        phone1 = request.form["phone1"]
        phone2 = request.form["phone2"]
        address = request.form["address"]

        # VALIDATION
        if not name or not phone1 or not phone2:
            flash("All required fields must be filled", "danger")
            return redirect(url_for("add_customer"))

        if not validate_phone(phone1) or not validate_phone(phone2):
            flash("Phone numbers must be valid (07XXXXXXXX)", "danger")
            return redirect(url_for("add_customer"))

        try:
            conn = get_db()
            conn.execute("""
                INSERT INTO customers (name, store_name, phone1, phone2, address, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (name, store, phone1, phone2, address, datetime.now()))

            conn.commit()
            conn.close()

            flash("Customer added successfully", "success")

        except sqlite3.IntegrityError:
            flash("Customer name already exists", "danger")

        except Exception as e:
            flash(f"Error: {str(e)}", "danger")

        return redirect(url_for("customers"))

    return render_template("add_customer.html", mode="add", customer=None)

@app.route("/edit_customer/<int:id>", methods=["GET", "POST"])
def edit_customer(id):
    conn = get_db()
    cur = conn.cursor()

    customer = cur.execute(
        "SELECT * FROM customers WHERE id=?", (id,)
    ).fetchone()

    if request.method == "POST":
        name = request.form["name"].strip()
        store = request.form["store"]
        phone1 = request.form["phone1"]
        phone2 = request.form["phone2"]
        address = request.form["address"]

        if not validate_phone(phone1) or not validate_phone(phone2):
            flash("Invalid phone format", "danger")
            return redirect(url_for("edit_customer", id=id))

        try:
            conn.execute("""
                UPDATE customers
                SET name=?, store_name=?, phone1=?, phone2=?, address=?
                WHERE id=?
            """, (name, store, phone1, phone2, address, id))

            conn.commit()
            flash("Customer updated", "success")

        except:
            flash("Duplicate name not allowed", "danger")

        return redirect(url_for("customers"))

    conn.close()
    return render_template("add_customer.html", mode="edit", customer=customer)


@app.route("/delete_customer/<int:id>")
def delete_customer(id):
    # This route checks for outstanding data and returns JSON for the modal
    conn = get_db()
    cur = conn.cursor()

    # Get customer details
    customer = cur.execute(
        "SELECT * FROM customers WHERE id=?", (id,)
    ).fetchone()

    if not customer:
        conn.close()
        flash("Customer not found", "danger")
        return redirect(url_for("customers"))

    # Check for outstanding credits (unsettled)
    outstanding_credits = cur.execute("""
        SELECT
            t.id,
            t.amount,
            t.date,
            t.due_date,
            t.payment_reference,
            COALESCE(SUM(s.amount), 0) as settled_amount,
            (t.amount - COALESCE(SUM(s.amount), 0)) as balance
        FROM transactions t
        LEFT JOIN settlements s ON s.credit_id = t.id
        WHERE t.customer_id = ? AND t.type = 'credit'
        GROUP BY t.id
        HAVING balance > 0
        ORDER BY t.date ASC
    """, (id,)).fetchall()

    # Check for overpayment
    overpayment_row = cur.execute("""
        SELECT overpayment FROM customer_balance WHERE customer_id = ?
    """, (id,)).fetchone()

    overpayment = float(overpayment_row["overpayment"]) if overpayment_row else 0

    # Check for any debits (payments) that might have been made
    debits = cur.execute("""
        SELECT COUNT(*) as count, SUM(amount) as total
        FROM transactions
        WHERE customer_id = ? AND type = 'debit'
    """, (id,)).fetchone()

    conn.close()

    # Prepare data for modal
    if outstanding_credits or overpayment > 0:
        credits_list = []
        for credit in outstanding_credits:
            credits_list.append({
                'id': credit['id'],
                'amount': float(credit['amount']),
                'balance': float(credit['balance']),
                'date': credit['date'],
                'due_date': credit['due_date'],
                'reference': credit['payment_reference'] or 'N/A'
            })

        return jsonify({
            'has_outstanding': True,
            'customer_id': id,
            'customer_name': customer['name'],
            'customer_store': customer['store_name'] or '',
            'outstanding_credits': credits_list,
            'overpayment': overpayment,
            'total_credits_count': len(credits_list),
            'total_outstanding_amount': sum([c['balance'] for c in credits_list]),
            'has_debits': debits['count'] > 0 if debits else False
        })

    # No outstanding data - delete directly
    return jsonify({
        'has_outstanding': False,
        'customer_id': id
    })


@app.route('/force_delete_customer', methods=['POST'])
def force_delete_customer():
    data = request.get_json()
    customer_id = data.get('customer_id')

    conn = get_db()
    cur = conn.cursor()

    try:
        # Get customer name for flash message
        customer = cur.execute(
            "SELECT name FROM customers WHERE id=?", (customer_id,)
        ).fetchone()

        if not customer:
            return jsonify({
                'success': False,
                'message': 'Customer not found'
            })

        # Delete from settlements (through transactions)
        cur.execute("""
            DELETE FROM settlements
            WHERE credit_id IN (SELECT id FROM transactions WHERE customer_id=?)
            OR debit_id IN (SELECT id FROM transactions WHERE customer_id=?)
        """, (customer_id, customer_id))

        # Delete from customer_balance
        cur.execute("""
            DELETE FROM customer_balance WHERE customer_id=?
        """, (customer_id,))

        # Delete all transactions
        cur.execute("""
            DELETE FROM transactions WHERE customer_id=?
        """, (customer_id,))

        # Delete the customer
        cur.execute("""
            DELETE FROM customers WHERE id=?
        """, (customer_id,))

        conn.commit()

        return jsonify({
            'success': True,
            'message': f'Customer "{customer["name"]}" and ALL associated data have been deleted successfully'
        })

    except Exception as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Delete Failed: {str(e)}'
        })
    finally:
        conn.close()

# =========================
# TRANSACTION
# =========================
@app.route("/transactions")
def transactions():
    conn = get_db()
    cur = conn.cursor()

    search = request.args.get("search", "").lower()
    ttype = request.args.get("type", "")

    rows = cur.execute("""
        SELECT t.*, c.name
        FROM transactions t
        JOIN customers c ON t.customer_id = c.id
        ORDER BY t.date DESC
    """).fetchall()

    data = []
    total_credit = 0
    total_debit = 0

    for r in rows:

        # FILTER: search
        if search and search not in r["name"].lower():
            continue

        # FILTER: type
        if ttype and r["type"] != ttype:
            continue

        data.append(r)

        # CALCULATE TOTALS
        if r["type"] == "credit":
            total_credit += r["amount"]
        else:
            total_debit += r["amount"]

    total_balance = total_credit - total_debit

    conn.close()

    return render_template(
        "transactions.html",
        data=data,
        total_credit=round(total_credit, 2),
        total_debit=round(total_debit, 2),
        total_balance=round(total_balance, 2)
    )


# =========================
# ADD TRANSACTION (CORE LOGIC)
# =========================

@app.route("/add_transaction", methods=["GET", "POST"])
def add_transaction():

    conn = get_db()
    cur = conn.cursor()

    customers = cur.execute("""
        SELECT *
        FROM customers
        ORDER BY name
    """).fetchall()

    # =====================================================
    # POST
    # =====================================================
    if request.method == "POST":

        cid = int(request.form["customer"])

        ttype = request.form["type"]

        amount = float(
            request.form["amount"]
            .replace(",", "")
            .strip()
        )

        date = request.form["date"]

        note = request.form.get("note", "")

        days = request.form.get("days", "")

        payment_method = request.form.get(
            "payment_method"
        )

        payment_reference = request.form.get(
            "payment_reference"
        )

        # =================================================
        # BASIC VALIDATION
        # =================================================
        if amount <= 0:

            flash(
                "Amount must be greater than 0",
                "danger"
            )

            conn.close()

            return redirect(
                url_for("add_transaction")
            )

        # =================================================
        # DEBIT REQUIRES PAYMENT METHOD
        # =================================================
        if ttype == "debit":

            if not payment_method:

                flash(
                    "Payment method required",
                    "danger"
                )

                conn.close()

                return redirect(
                    url_for("add_transaction")
                )

        # =================================================
        # CREDIT ENTRY
        # =================================================
        if ttype == "credit":

            due_date = None

            if days:

                due_date = (
                    datetime.strptime(
                        date,
                        "%Y-%m-%d"
                    ) + timedelta(days=int(days))
                ).strftime("%Y-%m-%d")

            # =============================================
            # SAVE CREDIT
            # =============================================
            cur.execute("""
                INSERT INTO transactions
                (
                    customer_id,
                    type,
                    amount,
                    date,
                    due_date,
                    note,
                    payment_method,
                    payment_reference
                )
                VALUES
                (
                    ?,
                    'credit',
                    ?,
                    ?,
                    ?,
                    ?,
                    NULL,
                    ?
                )
            """, (
                cid,
                amount,
                date,
                due_date,
                note,
                payment_reference
            ))

            new_credit_id = cur.lastrowid

            # =============================================
            # ENSURE BALANCE ROW
            # =============================================
            cur.execute("""
                INSERT OR IGNORE INTO customer_balance
                (
                    customer_id,
                    overpayment
                )
                VALUES (?, 0)
            """, (cid,))

            overpayment = get_overpayment(
                conn,
                cid
            )

            # =============================================
            # AUTO SETTLE USING OVERPAYMENT
            # =============================================
            if overpayment > 0:

                used = min(
                    overpayment,
                    amount
                )

                cur.execute("""
                    INSERT INTO settlements
                    (
                        credit_id,
                        debit_id,
                        amount
                    )
                    VALUES (?, NULL, ?)
                """, (
                    new_credit_id,
                    used
                ))

                cur.execute("""
                    UPDATE customer_balance
                    SET overpayment = overpayment - ?
                    WHERE customer_id = ?
                """, (
                    used,
                    cid
                ))

                flash(
                    f"Customer has Rs.{overpayment:.2f} "
                    f"overpayment. "
                    f"Rs.{used:.2f} was automatically settled.",
                    "info"
                )

        # =================================================
        # DEBIT ENTRY
        # =================================================
        else:

            selected_credits = request.form.getlist(
                "credit_ids"
            )

            # =============================================
            # MUST SELECT AT LEAST ONE
            # =============================================
            if not selected_credits:

                flash(
                    "Please select at least one credit",
                    "danger"
                )

                conn.close()

                return redirect(
                    url_for(
                        "add_transaction",
                        customer=cid
                    )
                )

            # =============================================
            # CALCULATE SELECTED TOTAL
            # =============================================
            selected_total = 0

            for credit_id in selected_credits:

                credit = cur.execute("""
                    SELECT amount
                    FROM transactions
                    WHERE id = ?
                    AND type = 'credit'
                """, (
                    credit_id,
                )).fetchone()

                if not credit:
                    continue

                settled = cur.execute("""
                    SELECT COALESCE(SUM(amount),0)
                    FROM settlements
                    WHERE credit_id = ?
                """, (
                    credit_id,
                )).fetchone()[0]

                available = (
                    credit["amount"] - settled
                )

                if available > 0:
                    selected_total += available

            # =============================================
            # GET TOTAL AVAILABLE CREDIT
            # =============================================
            result = cur.execute("""
                SELECT COALESCE(SUM(balance),0)
                FROM (
                    SELECT
                        t.amount -
                        COALESCE(SUM(s.amount),0)
                        AS balance

                    FROM transactions t

                    LEFT JOIN settlements s
                        ON s.credit_id = t.id

                    WHERE t.customer_id = ?
                    AND t.type = 'credit'

                    GROUP BY t.id

                    HAVING balance > 0
                )
            """, (
                cid,
            )).fetchone()

            total_available_credit = 0

            if result and result[0]:
                total_available_credit = result[0]

            # =============================================
            # IF FULL PAYMENT POSSIBLE
            # USER MUST SELECT ALL CREDITS
            # =============================================
            if amount >= total_available_credit:

                all_credits = cur.execute("""
                    SELECT COUNT(*)

                    FROM (
                        SELECT t.id

                        FROM transactions t

                        LEFT JOIN settlements s
                            ON s.credit_id = t.id

                        WHERE t.customer_id = ?
                        AND t.type = 'credit'

                        GROUP BY t.id

                        HAVING
                        (
                            t.amount -
                            COALESCE(SUM(s.amount),0)
                        ) > 0
                    )
                """, (
                    cid,
                )).fetchone()[0]

                if len(selected_credits) != all_credits:

                    flash(
                        "Please select all credits "
                        "because this payment can "
                        "fully settle all balances.",
                        "danger"
                    )

                    conn.close()

                    return redirect(
                        url_for(
                            "add_transaction",
                            customer=cid
                        )
                    )

            # =============================================
            # PARTIAL VALIDATION
            # =============================================
            else:

                if selected_total <= 0:

                    flash(
                        "Selected credits do not "
                        "contain remaining balances.",
                        "danger"
                    )

                    conn.close()

                    return redirect(
                        url_for(
                            "add_transaction",
                            customer=cid
                        )
                    )

            # =============================================
            # SAVE DEBIT
            # =============================================
            cur.execute("""
                INSERT INTO transactions
                (
                    customer_id,
                    type,
                    amount,
                    date,
                    note,
                    payment_method,
                    payment_reference
                )
                VALUES
                (
                    ?,
                    'debit',
                    ?,
                    ?,
                    ?,
                    ?,
                    ?
                )
            """, (
                cid,
                amount,
                date,
                note,
                payment_method,
                payment_reference
            ))

            debit_id = cur.lastrowid

            remaining = amount

            # =============================================
            # APPLY SETTLEMENTS
            # =============================================
            for credit_id in selected_credits:

                if remaining <= 0:
                    break

                credit = cur.execute("""
                    SELECT amount
                    FROM transactions
                    WHERE id = ?
                    AND type = 'credit'
                """, (
                    credit_id,
                )).fetchone()

                if not credit:
                    continue

                settled = cur.execute("""
                    SELECT COALESCE(SUM(amount),0)
                    FROM settlements
                    WHERE credit_id = ?
                """, (
                    credit_id,
                )).fetchone()[0]

                available = (
                    credit["amount"] - settled
                )

                if available <= 0:
                    continue

                settle_amount = min(
                    available,
                    remaining
                )

                cur.execute("""
                    INSERT INTO settlements
                    (
                        credit_id,
                        debit_id,
                        amount
                    )
                    VALUES (?, ?, ?)
                """, (
                    credit_id,
                    debit_id,
                    settle_amount
                ))

                remaining -= settle_amount

            # =============================================
            # HANDLE OVERPAYMENT
            # =============================================
            if remaining > 0:

                cur.execute("""
                    INSERT OR IGNORE INTO customer_balance
                    (
                        customer_id,
                        overpayment
                    )
                    VALUES (?, 0)
                """, (
                    cid,
                ))

                cur.execute("""
                    UPDATE customer_balance
                    SET overpayment =
                        overpayment + ?
                    WHERE customer_id = ?
                """, (
                    remaining,
                    cid
                ))

        # =================================================
        # SAVE CHANGES
        # =================================================
        conn.commit()

        conn.close()

        flash(
            "Transaction saved successfully",
            "success"
        )

        return redirect(
            url_for("transactions")
        )

    # =====================================================
    # GET REQUEST
    # =====================================================
    selected_customer = request.args.get(
        "customer"
    )

    credits = []

    if selected_customer:

        credits = cur.execute("""
            SELECT
                t.id,
                t.amount,
                t.date,
                c.name,

                COALESCE(
                    SUM(s.amount),
                    0
                ) AS settled,

                t.amount -
                COALESCE(
                    SUM(s.amount),
                    0
                ) AS balance

            FROM transactions t

            JOIN customers c
                ON c.id = t.customer_id

            LEFT JOIN settlements s
                ON s.credit_id = t.id

            WHERE t.type = 'credit'
            AND t.customer_id = ?

            GROUP BY t.id

            HAVING balance > 0

            ORDER BY t.date ASC
        """, (
            selected_customer,
        )).fetchall()

    conn.close()

    return render_template(
        "add_transaction.html",
        customers=customers,
        credits=credits,
        today=datetime.today().strftime(
            "%Y-%m-%d"
        )
    )


# =========================
# CHECK REFERENCE EXISTS
# =========================
@app.route("/check_reference")
def check_reference():

    ref = request.args.get("ref", "").strip()

    if not ref:
        return jsonify({
            "exists": False
        })

    conn = get_db()
    cur = conn.cursor()

    exists = cur.execute("""
        SELECT id
        FROM transactions
        WHERE LOWER(TRIM(payment_reference)) = LOWER(?)
    """, (ref,)).fetchone()

    conn.close()

    return jsonify({
        "exists": True if exists else False
    })

@app.route("/edit_transaction/<int:id>", methods=["GET", "POST"])
def edit_transaction(id):
    conn = get_db()
    cur = conn.cursor()

    # Get transaction
    trx = cur.execute("""
        SELECT * FROM transactions WHERE id=?
    """, (id,)).fetchone()

    if not trx:
        conn.close()
        flash("Transaction not found", "danger")
        return redirect(url_for("transactions"))

    customers = cur.execute("""
        SELECT * FROM customers ORDER BY name
    """).fetchall()

    # =========================
    # POST (UPDATE)
    # =========================
    if request.method == "POST":

        customer_id = int(request.form["customer"])
        ttype = request.form["type"]
        amount = float(request.form["amount"].replace(",", "").strip())
        date = request.form["date"]
        note = request.form.get("note", "")
        days = request.form.get("days", "")
        payment_method = request.form.get("payment_method")
        payment_reference = request.form.get("payment_reference")

        # =========================
        # VALIDATION
        # =========================
        if amount <= 0:
            flash("Amount must be greater than 0", "danger")
            conn.close()
            return redirect(url_for("edit_transaction", id=id))

        # Check if transaction type is changing
        type_changed = (trx["type"] != ttype)

        # =========================
        # TYPE CHANGE VALIDATION
        # =========================
        if type_changed:
            # If changing from CREDIT to DEBIT
            if trx["type"] == "credit" and ttype == "debit":
                # Check if this credit has any settlements
                settled = cur.execute("""
                    SELECT COALESCE(SUM(amount), 0) FROM settlements WHERE credit_id=?
                """, (id,)).fetchone()[0]

                if settled > 0:
                    flash(f"Cannot change Credit to Debit because Rs.{settled:.2f} has already been settled against this invoice!", "danger")
                    conn.close()
                    return redirect(url_for("edit_transaction", id=id))

            # If changing from DEBIT to CREDIT
            elif trx["type"] == "debit" and ttype == "credit":
                # Check if this debit has any settlements
                settled = cur.execute("""
                    SELECT COALESCE(SUM(amount), 0) FROM settlements WHERE debit_id=?
                """, (id,)).fetchone()[0]

                if settled > 0:
                    flash(f"Cannot change Debit to Credit because this payment of Rs.{settled:.2f} has already been used to settle invoices!", "danger")
                    conn.close()
                    return redirect(url_for("edit_transaction", id=id))

        # =========================
        # REFERENCE VALIDATION
        # =========================
        if payment_reference and payment_reference.strip():
            existing_ref = cur.execute("""
                SELECT id FROM transactions
                WHERE LOWER(TRIM(payment_reference)) = LOWER(?)
                AND id != ?
            """, (payment_reference.strip(), id)).fetchone()

            if existing_ref:
                flash("Reference already exists. Please use a unique reference.", "danger")
                conn.close()
                return redirect(url_for("edit_transaction", id=id))

        # =========================
        # CREDIT VALIDATION (Keep original type check)
        # =========================
        if trx["type"] == "credit" and not type_changed:
            settled = cur.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM settlements WHERE credit_id=?
            """, (id,)).fetchone()[0]

            if amount < settled:
                flash(f"Cannot reduce amount below settled amount (Rs.{settled:.2f})", "danger")
                conn.close()
                return redirect(url_for("edit_transaction", id=id))

        # =========================
        # DEBIT VALIDATION (Keep original type check)
        # =========================
        if trx["type"] == "debit" and not type_changed:
            settled = cur.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM settlements WHERE debit_id=?
            """, (id,)).fetchone()[0]

            if amount < settled:
                flash(f"Cannot reduce amount below settled amount (Rs.{settled:.2f})", "danger")
                conn.close()
                return redirect(url_for("edit_transaction", id=id))

        # =========================
        # DUE DATE CALCULATION
        # =========================
        due_date = None
        if ttype == "credit" and days:
            due_date = (
                datetime.strptime(date, "%Y-%m-%d") +
                timedelta(days=int(days))
            ).strftime("%Y-%m-%d")

        # =========================
        # UPDATE TRANSACTION
        # =========================
        cur.execute("""
            UPDATE transactions
            SET customer_id=?,
                type=?,
                amount=?,
                date=?,
                due_date=?,
                note=?,
                payment_method=?,
                payment_reference=?
            WHERE id=?
        """, (customer_id, ttype, amount, date, due_date, note,
              payment_method if payment_method else None,
              payment_reference if payment_reference else None, id))

        conn.commit()

        # =========================
        # REBUILD IF TYPE CHANGED OR CUSTOMER CHANGED
        # =========================
        if type_changed or customer_id != trx["customer_id"]:
            # Rebuild old customer if customer changed
            if customer_id != trx["customer_id"]:
                recalculate_customer(trx["customer_id"])
            # Rebuild new customer
            recalculate_customer(customer_id)
        else:
            # Just recalculate current customer
            recalculate_customer(customer_id)

        conn.close()

        flash("Transaction updated successfully", "success")
        return redirect(url_for("transactions"))

    # =========================
    # GET (LOAD FORM)
    # =========================
    days = ""
    if trx["type"] == "credit" and trx["due_date"]:
        d1 = datetime.strptime(trx["date"], "%Y-%m-%d")
        d2 = datetime.strptime(trx["due_date"], "%Y-%m-%d")
        days = (d2 - d1).days

    conn.close()

    return render_template(
        "edit_transaction.html",
        t=trx,
        customers=customers,
        days=days
    )

@app.route("/get_overpayment/<int:customer_id>")
def get_overpayment_api(customer_id):
    conn = get_db()
    overpayment = get_overpayment(conn, customer_id)
    conn.close()

    return {"overpayment": overpayment}


@app.route("/delete_transaction/<int:id>")
def delete_transaction(id):
    conn = get_db()
    cur = conn.cursor()

    # Get transaction details
    trx = cur.execute("""
        SELECT * FROM transactions WHERE id=?
    """, (id,)).fetchone()

    if not trx:
        conn.close()
        flash("Transaction not found", "danger")
        return redirect(url_for("transactions"))

    customer_id = trx["customer_id"]
    transaction_type = trx["type"]

    try:
        # Delete settlements related to this transaction
        if transaction_type == "credit":
            # Delete settlements where this credit is used
            cur.execute("DELETE FROM settlements WHERE credit_id=?", (id,))
        else:
            # Delete settlements where this debit is used
            cur.execute("DELETE FROM settlements WHERE debit_id=?", (id,))

        # Delete the transaction
        cur.execute("DELETE FROM transactions WHERE id=?", (id,))

        conn.commit()

        # Rebuild customer state
        recalculate_customer(customer_id)

        flash("Transaction deleted & balances updated successfully", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {str(e)}", "danger")
    finally:
        conn.close()

    return redirect(url_for("dashboard"))



# =========================
# LEDGER
# =========================

# =========================
# LEDGER PAGE (MODIFIED WITH CLEAR BUTTON CONDITION)
# =========================
@app.route("/ledger/<int:id>")
def ledger(id):
    conn = get_db()

    # Customer details
    customer = conn.execute(
        "SELECT * FROM customers WHERE id=?", (id,)
    ).fetchone()

    # Get all transactions
    transactions = conn.execute("""
        SELECT id, date, amount, type, note, payment_method, payment_reference
        FROM transactions
        WHERE customer_id=?
        ORDER BY date ASC, id ASC
    """, (id,)).fetchall()

    # Check for overpayment
    overpayment_row = conn.execute("""
        SELECT overpayment FROM customer_balance WHERE customer_id=?
    """, (id,)).fetchone()
    overpayment = float(overpayment_row["overpayment"]) if overpayment_row else 0

    # Check for unsettled credits
    unsettled_credits = conn.execute("""
        SELECT COUNT(*) as count
        FROM (
            SELECT
                t.id,
                t.amount,
                COALESCE(SUM(s.amount), 0) as settled
            FROM transactions t
            LEFT JOIN settlements s ON s.credit_id = t.id
            WHERE t.customer_id = ? AND t.type = 'credit'
            GROUP BY t.id
            HAVING (t.amount - COALESCE(SUM(s.amount), 0)) > 0
        )
    """, (id,)).fetchone()

    # Build ledger data
    all_rows = []
    for t in transactions:
        all_rows.append(dict(t))

    all_rows = sorted(all_rows, key=lambda x: (x["date"], x["id"]))

    balance = 0
    ledger_data = []
    total_credit = 0
    total_debit = 0

    for row in all_rows:
        if row["type"] == "credit":
            balance += row["amount"]
            total_credit += row["amount"]
        else:
            balance -= row["amount"]
            total_debit += row["amount"]

        ledger_data.append({
            "t": row,
            "balance": balance
        })

    conn.close()

    # Determine if clear button should be shown
    # Button එක පෙන්වන්නේ:
    # 1. Transactions තිබේ නම්
    # 2. Overpayment එකක් නැතිනම් (overpayment == 0)
    # 3. Unsettled credits නැතිනම් (unsettled_credits['count'] == 0)
    show_clear_button = (
        len(ledger_data) > 0 and
        overpayment == 0 and
        unsettled_credits['count'] == 0
    )

    return render_template(
        "ledger.html",
        customer=customer,
        data=ledger_data,
        total_credit=total_credit,
        total_debit=total_debit,
        show_clear_button=show_clear_button
    )


# =========================
# CLEAR ALL TRANSACTIONS FOR CUSTOMER
# =========================
@app.route("/clear_customer_transactions/<int:customer_id>", methods=["POST"])
def clear_customer_transactions(customer_id):
    conn = get_db()
    cur = conn.cursor()

    try:
        # Get customer details
        customer = cur.execute(
            "SELECT name FROM customers WHERE id=?", (customer_id,)
        ).fetchone()

        if not customer:
            return jsonify({
                'success': False,
                'message': 'Customer not found'
            })

        # Check for overpayment
        overpayment_row = cur.execute("""
            SELECT overpayment FROM customer_balance WHERE customer_id=?
        """, (customer_id,)).fetchone()
        overpayment = float(overpayment_row["overpayment"]) if overpayment_row else 0

        if overpayment > 0:
            return jsonify({
                'success': False,
                'message': 'Cannot clear transactions: Customer has overpayment. Please handle overpayment first.'
            })

        # Check for unsettled credits
        unsettled = cur.execute("""
            SELECT COUNT(*) as count
            FROM (
                SELECT
                    t.id,
                    t.amount,
                    COALESCE(SUM(s.amount), 0) as settled
                FROM transactions t
                LEFT JOIN settlements s ON s.credit_id = t.id
                WHERE t.customer_id = ? AND t.type = 'credit'
                GROUP BY t.id
                HAVING (t.amount - COALESCE(SUM(s.amount), 0)) > 0
            )
        """, (customer_id,)).fetchone()

        if unsettled['count'] > 0:
            return jsonify({
                'success': False,
                'message': f'Cannot clear transactions: Customer has {unsettled["count"]} unsettled credit(s). Please settle all credits first.'
            })

        # Check if there are any transactions
        transactions = cur.execute("""
            SELECT COUNT(*) as count FROM transactions WHERE customer_id=?
        """, (customer_id,)).fetchone()

        if transactions['count'] == 0:
            return jsonify({
                'success': False,
                'message': 'No transactions found for this customer'
            })

        # Delete from settlements
        cur.execute("""
            DELETE FROM settlements
            WHERE credit_id IN (SELECT id FROM transactions WHERE customer_id=?)
            OR debit_id IN (SELECT id FROM transactions WHERE customer_id=?)
        """, (customer_id, customer_id))

        # Delete all transactions
        cur.execute("""
            DELETE FROM transactions WHERE customer_id=?
        """, (customer_id,))

        conn.commit()

        return jsonify({
            'success': True,
            'message': f'All transactions for "{customer["name"]}" have been cleared successfully',
            'deleted_count': transactions['count']
        })

    except Exception as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Failed to clear transactions: {str(e)}'
        })
    finally:
        conn.close()


# =========================================================
# EXPORT ACCOUNTS RECEIVABLE REPORT
# =========================================================

@app.route("/export_receivable_report")
def export_receivable_report():

    try:

        # =====================================================
        # DATABASE
        # =====================================================
        conn = get_db()

        conn.row_factory = sqlite3.Row

        cur = conn.cursor()

        # =====================================================
        # CREATE WORKBOOK
        # =====================================================
        wb = Workbook()

        ws = wb.active

        ws.title = "Accounts Receivable Report"

        # =====================================================
        # SETTINGS
        # =====================================================
        START_ROW = 11

        MONEY_COLUMNS = [7, 8, 9, 10, 11, 12, 13]

        MONEY_FORMAT = '#,##0.00'

        # =====================================================
        # STYLES
        # =====================================================
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center = Alignment(
            horizontal='center',
            vertical='center'
        )

        left = Alignment(
            horizontal='left',
            vertical='center'
        )

        title_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        summary_fill = PatternFill(
            start_color="E2F0D9",
            end_color="E2F0D9",
            fill_type="solid"
        )

        total_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )

        # =====================================================
        # MAIN TITLE
        # =====================================================
        ws.merge_cells("A1:P2")

        ws["A1"] = "ACCOUNTS RECEIVABLE REPORT"

        ws["A1"].font = Font(
            bold=True,
            size=22,
            color="FFFFFF"
        )

        ws["A1"].alignment = center

        ws["A1"].fill = title_fill

        # =====================================================
        # COMPANY INFORMATION
        # =====================================================
        company_info = [

            ("Sales Officer 1", "Mr : Chamal"),

            ("Sales Officer 2", "Mr : Eranga"),

            ("Generated Date", datetime.today()),

            ("Area", "Wadiyamankada - HC"),

            ("Distributor", "Kishudha Expert Distribution")

        ]

        info_start = 4

        for i, (label, value) in enumerate(company_info):

            row_no = info_start + i

            # Label
            ws.merge_cells(f"A{row_no}:B{row_no}")
            ws[f"A{row_no}"] = label

            ws[f"A{row_no}"].font = Font(
                bold=True,
                color="FFFFFF"
            )

            ws[f"A{row_no}"].fill = title_fill

            ws[f"A{row_no}"].alignment = left

            ws[f"A{row_no}"].border = thin

            # Value
            ws.merge_cells(f"C{row_no}:D{row_no}")

            ws[f"C{row_no}"] = value

            ws[f"C{row_no}"].alignment = left

            if isinstance(value, datetime):
                ws[f"C{row_no}"].number_format = "DD-MMM-YYYY"

            for col in range(3, 5):

                cell = ws.cell(row_no, col)

                cell.border = thin

        # =====================================================
        # TABLE HEADERS
        # =====================================================
        headers = [

            "No",
            "Invoice Date",
            "Customer Name",
            "Invoice No",
            "Phone 1",
            "Phone 2",
            "Invoice Amount",
            "Paid Amount",
            "Outstanding Amount",
            "Below 14 Days",
            "14 - 21 Days",
            "21 - 30 Days",
            "Above 30 Days",
            "Today",
            "Days",
            "Remark"

        ]

        for col_num, header in enumerate(headers, start=1):

            cell = ws.cell(START_ROW, col_num)

            cell.value = header

            cell.font = Font(
                bold=True
            )

            cell.border = thin

            cell.alignment = center

            cell.fill = header_fill

        # =====================================================
        # LOAD DATABASE DATA
        # =====================================================
        rows = cur.execute("""
            SELECT
                t.id,
                t.date,
                t.amount,
                t.payment_reference,

                c.store_name,
                c.phone1,
                c.phone2,

                (
                    SELECT COALESCE(SUM(s.amount), 0)
                    FROM settlements s
                    WHERE s.credit_id = t.id
                ) AS paid

            FROM transactions t

            INNER JOIN customers c
                ON c.id = t.customer_id

            WHERE t.type = 'credit'

            ORDER BY t.date ASC
        """).fetchall()

        # =====================================================
        # INSERT DATA
        # =====================================================
        current_row = START_ROW + 1

        for index, row in enumerate(rows, start=1):

            invoice_amount = float(row["amount"] or 0)

            paid_amount = float(row["paid"] or 0)

            # =================================================
            # DATE
            # =================================================
            try:

                invoice_date = datetime.strptime(
                    str(row["date"]),
                    "%Y-%m-%d"
                )

            except:

                invoice_date = datetime.today()

            # =================================================
            # DATA
            # =================================================
            ws[f"A{current_row}"] = index

            ws[f"B{current_row}"] = invoice_date

            ws[f"B{current_row}"].number_format = "DD-MMM-YYYY"

            ws[f"C{current_row}"] = row["store_name"] or ""

            ws[f"D{current_row}"] = row["payment_reference"] or ""

            ws[f"E{current_row}"] = row["phone1"] or ""

            ws[f"F{current_row}"] = row["phone2"] or "-"

            ws[f"G{current_row}"] = invoice_amount

            ws[f"H{current_row}"] = paid_amount

            # =================================================
            # BALANCE
            # =================================================
            ws[f"I{current_row}"] = (
                f"=G{current_row}-H{current_row}"
            )

            # =================================================
            # AGEING
            # =================================================
            ws[f"J{current_row}"] = (
                f'=IF(TODAY()-B{current_row}<=14,I{current_row},0)'
            )

            ws[f"K{current_row}"] = (
                f'=IF(AND(TODAY()-B{current_row}>14,'
                f'TODAY()-B{current_row}<=21),'
                f'I{current_row},0)'
            )

            ws[f"L{current_row}"] = (
                f'=IF(AND(TODAY()-B{current_row}>21,'
                f'TODAY()-B{current_row}<=30),'
                f'I{current_row},0)'
            )

            ws[f"M{current_row}"] = (
                f'=IF(TODAY()-B{current_row}>30,'
                f'I{current_row},0)'
            )

            # =================================================
            # TODAY DATE
            # =================================================
            ws[f"N{current_row}"] = datetime.today()

            ws[f"N{current_row}"].number_format = "DD-MMM-YYYY"

            # =================================================
            # DAYS
            # =================================================
            ws[f"O{current_row}"] = (
                f'=INT(N{current_row}-B{current_row})'
            )

            ws[f"O{current_row}"].number_format = "0"

            # =================================================
            # REMARK
            # =================================================
            ws[f"P{current_row}"] = ""

            # =================================================
            # STYLE
            # =================================================
            for col in range(1, 17):

                cell = ws.cell(current_row, col)

                cell.border = thin

                if col in [3, 16]:
                    cell.alignment = left
                else:
                    cell.alignment = center

                if col in MONEY_COLUMNS:
                    cell.number_format = MONEY_FORMAT

            current_row += 1

        # =====================================================
        # TOTAL ROW
        # =====================================================
        total_row = current_row

        ws[f"A{total_row}"] = "TOTAL"

        ws[f"G{total_row}"] = (
            f"=SUM(G{START_ROW+1}:G{total_row-1})"
        )

        ws[f"H{total_row}"] = (
            f"=SUM(H{START_ROW+1}:H{total_row-1})"
        )

        ws[f"I{total_row}"] = (
            f"=SUM(I{START_ROW+1}:I{total_row-1})"
        )

        ws[f"J{total_row}"] = (
            f"=SUM(J{START_ROW+1}:J{total_row-1})"
        )

        ws[f"K{total_row}"] = (
            f"=SUM(K{START_ROW+1}:K{total_row-1})"
        )

        ws[f"L{total_row}"] = (
            f"=SUM(L{START_ROW+1}:L{total_row-1})"
        )

        ws[f"M{total_row}"] = (
            f"=SUM(M{START_ROW+1}:M{total_row-1})"
        )

        # =====================================================
        # TOTAL ROW STYLE
        # =====================================================
        for col in range(1, 17):

            cell = ws.cell(total_row, col)

            cell.font = Font(
                bold=True,
                size=11
            )

            cell.fill = total_fill

            cell.border = thin

            cell.alignment = center

            if col in MONEY_COLUMNS:
                cell.number_format = MONEY_FORMAT

        # =====================================================
        # SUMMARY SECTION
        # =====================================================
        summary_row = total_row + 4

        ws.merge_cells(f"A{summary_row}:C{summary_row}")

        ws[f"A{summary_row}"] = "SUMMARY REPORT"

        ws[f"A{summary_row}"].font = Font(
            bold=True,
            size=16,
            color="FFFFFF"
        )

        ws[f"A{summary_row}"].fill = title_fill

        ws[f"A{summary_row}"].alignment = center

        summary_data = [

            (
                "Total Credit Bill",
                f"=SUM(G{START_ROW+1}:G{total_row-1})"
            ),

            (
                "Cash Received",
                f"=SUM(H{START_ROW+1}:H{total_row-1})"
            ),

            (
                "Outstanding Amount",
                f"=SUM(I{START_ROW+1}:I{total_row-1})"
            ),

            (
                "Below 14 Days",
                f"=SUM(J{START_ROW+1}:J{total_row-1})"
            ),

            (
                "14 - 21 Days",
                f"=SUM(K{START_ROW+1}:K{total_row-1})"
            ),

            (
                "21 - 30 Days",
                f"=SUM(L{START_ROW+1}:L{total_row-1})"
            ),

            (
                "Above 30 Days",
                f"=SUM(M{START_ROW+1}:M{total_row-1})"
            )

        ]

        for i, (label, formula) in enumerate(summary_data, start=1):

            row_no = summary_row + i

            # Label
            ws.merge_cells(f"A{row_no}:B{row_no}")
            ws[f"A{row_no}"] = label

            ws[f"A{row_no}"].font = Font(
                bold=True
            )

            ws[f"A{row_no}"].fill = summary_fill

            ws[f"A{row_no}"].alignment = left

            # ws[f"A{row_no}"].border = thin
            # ws[f"B{row_no}"].border = thin

            for col in range(2, 3):

                cell = ws.cell(row_no, col)

                cell.border = thin

            # Formula

            ws[f"C{row_no}"] = formula

            ws[f"C{row_no}"].font = Font(
                bold=True
            )

            ws[f"C{row_no}"].number_format = MONEY_FORMAT

            ws[f"C{row_no}"].alignment = center

            ws[f"C{row_no}"].border = thin

        # =====================================================
        # COLUMN WIDTHS
        # =====================================================
        widths = {

            "A": 10,
            "B": 16,
            "C": 35,
            "D": 22,
            "E": 15,
            "F": 15,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 18,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 16,
            "O": 10,
            "P": 25

        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # =====================================================
        # FREEZE HEADER
        # =====================================================
        ws.freeze_panes = "A12"

        # =====================================================
        # SAVE FILE
        # =====================================================
        output = BytesIO()

        wb.save(output)

        output.seek(0)

        conn.close()

        # =====================================================
        # DOWNLOAD
        # =====================================================
        return send_file(
            output,
            as_attachment=True,
            download_name="Accounts_Receivable_Report.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

    except Exception as e:

        return f"Error generating report : {str(e)}"


# ---------------- AUTO OPEN BROWSER ----------------
def open_browser():
    webbrowser.open_new("http://127.0.0.1:5000/")


# ---------------- MAIN ----------------
if __name__ == '__main__':


    # Prevent duplicate browser opening in debug mode
    if not os.environ.get("WERKZEUG_RUN_MAIN"):
        threading.Timer(1.5, open_browser).start()

    app.run(host="127.0.0.1", port=5000, debug=False)
